import streamlit as st
import pandas as pd
from datetime import datetime
import json
import os
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Importación de Módulos Externos ---
from auditor_module import auditor_deliverables_portal
from admin_uploader_module import admin_file_uploader

# --- Configuración Inicial de la Página (Debe ser la primera orden de Streamlit) ---
st.set_page_config(
    page_title="Portal de Contribuyente - RI Consultores",
    page_icon="📊",
    layout="wide"
)

# --- Configuración de Persistencia en Disco ---
DB_FILE = "submissions_db.json"
EMPLOYEES_FILE = "employees_db.json"
EVENTUALES_FILE = "eventuales_db.json"
UPLOAD_DIR = "uploaded_files"

if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

def load_submissions():
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_submission_to_disk(submission_data):
    submissions = load_submissions()
    submissions.append(submission_data)
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(submissions, f, ensure_ascii=False, indent=4)

def load_json_db(file_path):
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_json_db(file_path, data):
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def save_files_to_folder(file_list, client_name, periodo_str, category):
    saved_files_info = []
    if not file_list:
        return saved_files_info
        
    safe_client = client_name.replace(" ", "_").replace(".", "")
    safe_periodo = periodo_str.replace(" ", "_")
    folder_path = os.path.join(UPLOAD_DIR, safe_client, safe_periodo, category)
    os.makedirs(folder_path, exist_ok=True)
    
    for file_obj in file_list:
        file_path = os.path.join(folder_path, file_obj.name)
        file_obj.seek(0)
        with open(file_path, "wb") as f:
            f.write(file_obj.getbuffer())
        saved_files_info.append({
            "name": file_obj.name,
            "path": file_path
        })
    return saved_files_info

def create_zip_buffer(json_list, pdf_list):
    """Crea un archivo ZIP en memoria con los JSONs y PDFs proporcionados."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file_info in (json_list or []):
            if os.path.exists(file_info['path']):
                zip_file.write(file_info['path'], arcname=file_info['name'])
        for file_info in (pdf_list or []):
            if os.path.exists(file_info['path']):
                zip_file.write(file_info['path'], arcname=file_info['name'])
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

# --- Funciones de Cálculo de Ley (El Salvador) ---
def calcular_isss_quincenal(sueldo_base_q):
    # ISSS: 3% con techo de $500 quincenal ($1,000 mensual)
    base = min(sueldo_base_q, 500.0)
    return round(base * 0.03, 2)

def calcular_afp_quincenal(sueldo_base_q):
    # AFP: 7.25% sobre el sueldo base quincenal
    return round(sueldo_base_q * 0.0725, 2)

def calcular_renta_quincenal(sueldo_neto_isss_afp):
    # Tramos de Renta Quincenal (El Salvador)
    b = sueldo_neto_isss_afp
    if b <= 236.00:
        return 0.0
    elif b <= 447.62:
        return round((b - 236.00) * 0.10 + 8.83, 2)
    elif b <= 1019.05:
        return round((b - 447.63) * 0.20 + 30.00, 2)
    else:
        return round((b - 1019.05) * 0.30 + 144.28, 2)

def extract_invoice_summary(file_list):
    summary_data = []
    if not file_list:
        return pd.DataFrame()
    
    for file_info in file_list:
        path = file_info.get("path")
        if path and os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8-sig") as f:
                    content = json.load(f)
                
                items = content if isinstance(content, list) else [content]
                for item in items:
                    doc_num = None
                    nc_root = item.get("numeroControl")
                    if nc_root and str(nc_root).startswith("DTE-03"):
                        doc_num = nc_root
                        
                    ident = item.get("identificacion", {})
                    if not doc_num and isinstance(ident, dict):
                        nc_ident = ident.get("numeroControl")
                        if nc_ident and str(nc_ident).startswith("DTE-03"):
                            doc_num = nc_ident
                                
                    if not doc_num:
                        doc_num = (
                            nc_root or 
                            (ident.get("numeroControl") if isinstance(ident, dict) else None) or
                            item.get("codigoGeneracion") or 
                            item.get("numDocumento") or 
                            file_info["name"]
                        )
                    
                    gen_code = None
                    if isinstance(ident, dict):
                        gen_code = ident.get("codigoGeneracion")
                    if not gen_code:
                        gen_code = item.get("codigoGeneracion") or item.get("selloRecibido") or "N/A"
                    
                    resumen = item.get("resumen", {})
                    if not isinstance(resumen, dict):
                        resumen = {}
                    
                    val = (
                        resumen.get("totalGravada") or 
                        resumen.get("subTotal") or 
                        resumen.get("subTotalVentas") or 
                        resumen.get("montoTotalOperacion") or 
                        item.get("totalGravada") or
                        item.get("subtotal") or 
                        item.get("valor") or 
                        0.0
                    )
                    
                    iva = (
                        resumen.get("totalIva") or 
                        resumen.get("iva") or 
                        resumen.get("ivaRenta") or 
                        resumen.get("ivaPerci1") or 
                        resumen.get("ivaRete1") or 
                        item.get("totalIva") or 
                        item.get("iva") or 
                        0.0
                    )
                    
                    if not iva and "tributos" in resumen and isinstance(resumen["tributos"], list):
                        iva_tributos = 0.0
                        for trib in resumen["tributos"]:
                            if isinstance(trib, dict):
                                val_trib = trib.get("valor") or trib.get("valTributo") or 0.0
                                try:
                                    iva_tributos += float(val_trib)
                                except:
                                    pass
                        if iva_tributos > 0:
                            iva = iva_tributos

                    total = (
                        resumen.get("totalPagar") or 
                        resumen.get("montoTotalOperacion") or 
                        resumen.get("total") or 
                        item.get("totalPagar") or 
                        item.get("total") or 
                        0.0
                    )
                    
                    try:
                        val_f = float(val) if val is not None else 0.0
                        iva_f = float(iva) if iva is not None else 0.0
                        total_f = float(total) if total is not None else 0.0
                        
                        if total_f == 0.0 and val_f > 0.0:
                            total_f = val_f + iva_f
                        elif val_f == 0.0 and total_f > 0.0 and iva_f > 0.0:
                            val_f = total_f - iva_f
                    except:
                        val_f, iva_f, total_f = 0.0, 0.0, 0.0
                        
                    summary_data.append({
                        "Código de Generación": str(gen_code),
                        "Número de Control": str(doc_num),
                        "Valor": val_f,
                        "IVA": iva_f,
                        "Total": total_f
                    })
            except Exception:
                summary_data.append({
                    "Código de Generación": "N/A",
                    "Número de Control": file_info["name"],
                    "Valor": 0.0,
                    "IVA": 0.0,
                    "Total": 0.0
                })
    return pd.DataFrame(summary_data)

# --- Inicialización de Estados de Sesión y Bases Persistentes ---
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "user_role" not in st.session_state:
    st.session_state.user_role = None
if "username" not in st.session_state:
    st.session_state.username = ""
if "user_id" not in st.session_state:
    st.session_state.user_id = ""

if "clients_db" not in st.session_state:
    st.session_state.clients_db = {}
if "employees_db" not in st.session_state:
    st.session_state.employees_db = load_json_db(EMPLOYEES_FILE)
if "eventuales_db" not in st.session_state:
    st.session_state.eventuales_db = load_json_db(EVENTUALES_FILE)

# --- Sincronización de Clientes Oficiales ---
official_clients = {
    "admin": {"password": "admin123", "role": "admin", "name": "Administrador General"},
    "soluciones_503": {"password": "sol503_2026", "role": "client", "name": "Soluciones 503 S.A.S. de C.V"},
    "distribuidora_libertad": {"password": "libertad_2026", "role": "client", "name": "Distribuidora Libertad"},
    "leftech": {"password": "leftech_2026", "role": "client", "name": "Leftech"},
    "cedillo": {"password": "cedillo_2026", "role": "client", "name": "Cedillo"},
    "mercadito_rosa": {"password": "rosa_2026", "role": "client", "name": "Mercadito Rosa de Saron AC"}
}

for k, v in official_clients.items():
    st.session_state.clients_db[k] = v

# --- Pantalla de Login ---
def login_screen():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🔐 RI Consultores")
        st.markdown("### Portal de Gestión Documental y Auditoría")
        
        with st.form("login_form"):
            username = st.text_input("Usuario")
            password = st.text_input("Contraseña", type="password")
            submit = st.form_submit_button("Iniciar Sesión", use_container_width=True)
            
            if submit:
                user_key = username.strip().lower()
                if user_key in st.session_state.clients_db and st.session_state.clients_db[user_key]["password"] == password:
                    st.session_state.logged_in = True
                    st.session_state.user_role = st.session_state.clients_db[user_key]["role"]
                    st.session_state.username = st.session_state.clients_db[user_key]["name"]
                    st.session_state.user_id = user_key
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos.")

# --- Panel de Administración ---
def admin_dashboard():
    st.title("🎛️ Panel de Control - Administrador")
    st.markdown("Supervisa el cumplimiento fiscal, administra cuentas, revisa los documentos cargados y gestiona entregables de auditoría.")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Estatus y Archivos Recibidos", "📤 Cargar Entregables de Auditoría", "➕ Crear Nuevo Usuario", "👥 Listado de Cuentas"])
    
    with tab1:
        st.subheader("Control de Recepción y Descarga de Documentos y Planillas")
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            filtro_mes = st.selectbox("Filtrar por Mes", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], index=5)
        with col_f2:
            filtro_anio = st.selectbox("Filtrar por Año", [2026, 2025], index=0)
            
        periodo_seleccionado = f"{filtro_mes} {filtro_anio}"
        all_submissions = load_submissions()
        envios_periodo = [s for s in all_submissions if s["periodo"] == periodo_seleccionado]
        
        if envios_periodo:
            st.success(f"Se encontraron {len(envios_periodo)} registros/entregas para el periodo {periodo_seleccionado}.")
            for idx, envio in enumerate(envios_periodo):
                if envio.get("type") == "planilla_quincenal_empleados":
                    with st.expander(f"💼 [PLANILLA QUINCENAL SUELDOS] {envio['client']} — ({envio.get('quincena', 'Quincena')}) — Entregado el {envio['fecha']}"):
                        st.markdown(f"**Periodo Fiscal:** {envio['periodo']} | **Quincena:** {envio.get('quincena', 'N/A')}")
                        st.markdown(f"**Resumen Consolidado:** Total Devengado: **${envio['total_devengado']:,.2f}** | ISSS: **${envio['total_isss']:,.2f}** | AFP: **${envio['total_afp']:,.2f}** | RENTA: **${envio['total_renta']:,.2f}** | Líquido a Pagar: **${envio['total_liquido']:,.2f}**")
                        
                        df_pq_admin = pd.DataFrame(envio['data'])
                        output_pq_adm = io.BytesIO()
                        total_row_pq_adm = pd.DataFrame([{
                            "Empleado": "TOTALES GENERALES",
                            "DUI": "",
                            "Código Renta": "",
                            "Salario Base M.": 0.0,
                            "Bonif. / Ingresos": 0.0,
                            "Total Devengado Q.": envio['total_devengado'],
                            "ISSS": envio['total_isss'],
                            "AFP": envio['total_afp'],
                            "RENTA": envio['total_renta'],
                            "Líquido a Pagar": envio['total_liquido']
                        }])
                        df_exp_pq_adm = pd.concat([df_pq_admin, total_row_pq_adm], ignore_index=True)
                        
                        with pd.ExcelWriter(output_pq_adm, engine='openpyxl') as writer:
                            df_exp_pq_adm.to_excel(writer, index=False, sheet_name='Planilla Quincenal', startrow=4)
                            workbook = writer.book
                            worksheet = writer.sheets['Planilla Quincenal']
                            
                            worksheet['A1'] = f"RI CONSULTORES — PLANILLA DE SUELDOS QUINCENAL ({envio.get('quincena', '').upper()})"
                            worksheet['A2'] = f"Periodo Fiscal: {envio['periodo']}"
                            worksheet['A3'] = f"Cliente / Contribuyente: {envio['client']}"
                            
                            title_font = Font(name='Calibri', size=12, bold=True, color='1F4E78')
                            subtitle_font = Font(name='Calibri', size=10, bold=True, color='595959')
                            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                            total_font = Font(name='Calibri', size=11, bold=True, color='000000')
                            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                            
                            align_center = Alignment(horizontal='center', vertical='center')
                            align_right = Alignment(horizontal='right', vertical='center')
                            align_left = Alignment(horizontal='left', vertical='center')
                            
                            thin_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
                            )
                            thick_bottom_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='double', color='000000')
                            )
                            
                            worksheet['A1'].font = title_font
                            worksheet['A2'].font = subtitle_font
                            worksheet['A3'].font = subtitle_font
                            
                            header_row_idx = 5
                            for col_num in range(1, len(df_exp_pq_adm.columns) + 1):
                                cell = worksheet.cell(row=header_row_idx, column=col_num)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = align_center
                                cell.border = thin_border
                                
                            total_row_idx = header_row_idx + len(df_exp_pq_adm)
                            for row_idx in range(header_row_idx + 1, total_row_idx + 1):
                                is_total_row = (row_idx == total_row_idx)
                                for col_idx in range(1, len(df_exp_pq_adm.columns) + 1):
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    if is_total_row:
                                        cell.font = total_font
                                        cell.fill = total_fill
                                        cell.border = thick_bottom_border
                                    else:
                                        cell.border = thin_border
                                        
                                    if col_idx >= 4:
                                        cell.number_format = '$#,##0.00'
                                        cell.alignment = align_right
                                    elif col_idx in [2, 3]:
                                        cell.alignment = align_center
                                    else:
                                        cell.alignment = align_left
                                        
                            for col in worksheet.columns:
                                max_len = 0
                                for cell in col:
                                    if cell.row >= 5:
                                        val_str = str(cell.value or '')
                                        if len(val_str) > max_len:
                                            max_len = len(val_str)
                                col_letter = get_column_letter(col[0].column)
                                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 18)
                                
                        excel_pq_adm_data = output_pq_adm.getvalue()
                        safe_c_name = envio['client'].replace(" ", "_").replace(".", "")
                        
                        st.download_button(
                            label="📥 Descargar Planilla Quincenal del Cliente (Excel)",
                            data=excel_pq_adm_data,
                            file_name=f"Planilla_Quincenal_{safe_c_name}_{envio['periodo'].replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_adm_pq_{idx}"
                        )

                elif envio.get("type") == "planilla_eventuales":
                    with st.expander(f"🧾 [PLANILLA EVENTUALES] {envio['client']} — Entregado el {envio['fecha']}"):
                        st.markdown(f"**Periodo Fiscal:** {envio['periodo']}")
                        st.markdown(f"**Resumen Consolidado:** Total Bruto: **${envio['total_bruto']:,.2f}** | Retención 10%: **${envio['total_retencion']:,.2f}** | Líquido a Pagar: **${envio['total_liquido']:,.2f}**")
                        
                        df_ev_admin = pd.DataFrame(envio['data'])
                        output_adm = io.BytesIO()
                        total_row_adm = pd.DataFrame([{
                            "Prestador Eventual": "TOTALES GENERALES",
                            "DUI": "",
                            "Monto Bruto": envio['total_bruto'],
                            "Retención 10%": envio['total_retencion'],
                            "Líquido a Pagar": envio['total_liquido']
                        }])
                        df_exp_adm = pd.concat([df_ev_admin, total_row_adm], ignore_index=True)
                        
                        with pd.ExcelWriter(output_adm, engine='openpyxl') as writer:
                            df_exp_adm.to_excel(writer, index=False, sheet_name='Planilla Eventuales', startrow=4)
                            workbook = writer.book
                            worksheet = writer.sheets['Planilla Eventuales']
                            
                            worksheet['A1'] = "RI CONSULTORES — PLANILLA DE RETENCIÓN DE RENTA (10% EVENTUALES)"
                            worksheet['A2'] = f"Periodo Fiscal / Mes de Referencia: {envio['periodo']}"
                            worksheet['A3'] = f"Cliente / Contribuyente: {envio['client']}"
                            
                            title_font = Font(name='Calibri', size=12, bold=True, color='1F4E78')
                            subtitle_font = Font(name='Calibri', size=10, bold=True, color='595959')
                            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                            total_font = Font(name='Calibri', size=11, bold=True, color='000000')
                            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                            
                            align_center = Alignment(horizontal='center', vertical='center')
                            align_right = Alignment(horizontal='right', vertical='center')
                            align_left = Alignment(horizontal='left', vertical='center')
                            
                            thin_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
                            )
                            thick_bottom_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='double', color='000000')
                            )
                            
                            worksheet['A1'].font = title_font
                            worksheet['A2'].font = subtitle_font
                            worksheet['A3'].font = subtitle_font
                            
                            header_row_idx = 5
                            for col_num in range(1, len(df_exp_adm.columns) + 1):
                                cell = worksheet.cell(row=header_row_idx, column=col_num)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = align_center
                                cell.border = thin_border
                                
                            total_row_idx = header_row_idx + len(df_exp_adm)
                            for row_idx in range(header_row_idx + 1, total_row_idx + 1):
                                is_total_row = (row_idx == total_row_idx)
                                for col_idx in range(1, len(df_exp_adm.columns) + 1):
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    if is_total_row:
                                        cell.font = total_font
                                        cell.fill = total_fill
                                        cell.border = thick_bottom_border
                                    else:
                                        cell.border = thin_border
                                        
                                    if col_idx in [3, 4, 5]:
                                        cell.number_format = '$#,##0.00'
                                        cell.alignment = align_right
                                    elif col_idx == 2:
                                        cell.alignment = align_center
                                    else:
                                        cell.alignment = align_left
                                        
                            for col in worksheet.columns:
                                max_len = 0
                                for cell in col:
                                    if cell.row >= 5:
                                        val_str = str(cell.value or '')
                                        if len(val_str) > max_len:
                                            max_len = len(val_str)
                                col_letter = get_column_letter(col[0].column)
                                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 18)
                                
                        excel_adm_data = output_adm.getvalue()
                        safe_c_name = envio['client'].replace(" ", "_").replace(".", "")
                        
                        st.download_button(
                            label="📥 Descargar Planilla de Eventuales del Cliente (Excel)",
                            data=excel_adm_data,
                            file_name=f"Planilla_Eventuales_10_{safe_c_name}_{envio['periodo'].replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_adm_ev_{idx}"
                        )
                else:
                    with st.expander(f"📁 {envio['client']} — Entregado el {envio['fecha']}"):
                        if envio.get('notes'):
                            st.info(f"**📝 Notas / Aclaraciones del Cliente:**\n\n{envio['notes']}")
                        
                        col_d1, col_d2 = st.columns(2)
                        with col_d1:
                            st.markdown("**📈 Ventas:**")
                            has_sales = envio.get('sales_json_list') or envio.get('sales_pdf_list')
                            if has_sales:
                                zip_sales_bytes = create_zip_buffer(envio.get('sales_json_list'), envio.get('sales_pdf_list'))
                                safe_client_name = envio['client'].replace(" ", "_").replace(".", "")
                                st.download_button(
                                    label="📦 Descargar todas las Ventas (ZIP)",
                                    data=zip_sales_bytes,
                                    file_name=f"Ventas_{safe_client_name}_{envio['periodo'].replace(' ', '_')}.zip",
                                    mime="application/zip",
                                    key=f"zip_sales_{idx}"
                                )
                            else:
                                st.text("Sin archivos de ventas")
                                
                        with col_d2:
                            st.markdown("**📉 Compras y Gastos:**")
                            has_purch = envio.get('purch_json_list') or envio.get('purch_pdf_list')
                            if has_purch:
                                zip_purch_bytes = create_zip_buffer(envio.get('purch_json_list'), envio.get('purch_pdf_list'))
                                safe_client_name = envio['client'].replace(" ", "_").replace(".", "")
                                st.download_button(
                                    label="📦 Descargar todas las Compras (ZIP)",
                                    data=zip_purch_bytes,
                                    file_name=f"Compras_{safe_client_name}_{envio['periodo'].replace(' ', '_')}.zip",
                                    mime="application/zip",
                                    key=f"zip_purch_{idx}"
                                )
                            else:
                                st.text("Sin archivos de compras")
        else:
            st.info(f"No hay documentos ni planillas registrados para el periodo {periodo_seleccionado} todavía.")

    with tab2:
        admin_file_uploader()

    with tab3:
        st.subheader("Dar de alta a un nuevo cliente")
        with st.form("new_client_form"):
            new_user_id = st.text_input("Identificador único de usuario (ej. empresa_abc)").strip().lower()
            company_name = st.text_input("Nombre Comercial / Razón Social")
            temp_pass = st.text_input("Contraseña Temporal", type="password")
            create_btn = st.form_submit_button("Registrar Cliente en el Sistema")
            
            if create_btn:
                if new_user_id and company_name and temp_pass:
                    if new_user_id in st.session_state.clients_db:
                        st.error("Ese identificador ya existe.")
                    else:
                        st.session_state.clients_db[new_user_id] = {
                            "password": temp_pass,
                            "role": "client",
                            "name": company_name
                        }
                        st.success(f"¡Cliente **{company_name}** registrado con éxito!")
                else:
                    st.warning("Completa todos los campos.")

    with tab4:
        st.subheader("Cuentas de Clientes Activas")
        client_accounts = [{"Usuario ID": k, "Nombre": v["name"]} for k, v in st.session_state.clients_db.items() if v["role"] == "client"]
        if client_accounts:
            st.dataframe(pd.DataFrame(client_accounts), use_container_width=True)
        else:
            st.warning("No hay clientes registrados.")

# --- Panel del Cliente con Pestañas y Portal de Auditoría ---
def client_dashboard():
    st.title(f"📁 PORTAL DE CONTRIBUYENTE — {st.session_state.username}")
    st.markdown("Gestión documental, notas aclaratorias, generación de planillas fiscales y portal de auditoría.")
    
    col_p1, col_p2, col_p3 = st.columns(3)
    with col_p1:
        mes = st.selectbox("MES FISCAL", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], index=5, key="client_mes")
    with col_p2:
        anio = st.selectbox("AÑO FISCAL", [2026, 2025], index=0, key="client_anio")
    with col_p3:
        quincena_op = st.selectbox("QUINCENA", ["Primera Quincena (Del 1 al 15)", "Segunda Quincena (Del 16 al Fin de Mes)"], key="client_quincena")
        
    periodo_str = f"{mes} {anio}"
    
    current_user_id = st.session_state.get("user_id", st.session_state.username)
    all_submissions = load_submissions()
    mis_envios = [s for s in all_submissions if s.get("user_id") == current_user_id or s.get("client") == st.session_state.username]
    envio_actual = next((s for s in mis_envios if s["periodo"] == periodo_str and s.get("type") not in ["planilla_eventuales", "planilla_quincenal_empleados"]), None)
    
    st.markdown("### 📌 Estatus y Acciones Requeridas")
    if envio_actual:
        st.success(f"✔️ **Periodo {periodo_str} al día:** Tus documentos han sido recibidos correctamente y se encuentran en proceso de auditoría por RI Consultores.")
    else:
        st.warning(f"⚠️ **Acción requerida para {periodo_str}:** Aún no has enviado la información de tus documentos de Ventas y Compras. Por favor cárgalos en la pestaña correspondiente.")
        
    st.divider()
    
    # --- PESTAÑAS PRINCIPALES DEL CLIENTE ---
    client_tab1, client_tab2, client_tab3, client_tab4, client_tab5 = st.tabs([
        "📁 CARGA DOCUMENTAL Y NOTAS", 
        "💼 GENERADOR DE PLANILLAS", 
        "📊 HISTORIAL Y RESUMEN",
        "🧾 BASE Y PLANILLA DE EVENTUALES (10%)",
        "🔍 ENTREGABLES DE AUDITORÍA"
    ])
    
    with client_tab1:
        with st.form("upload_form"):
            col_v, col_c = st.columns(2)
            
            with col_v:
                st.subheader("📈 Ventas")
                sales_json = st.file_uploader("Arrastra tus JSON de Ventas (Múltiples)", type=["json"], accept_multiple_files=True, key="s_json")
                sales_pdf = st.file_uploader("Arrastra tus PDFs/ZIP de Ventas (Múltiples)", type=["pdf", "zip"], accept_multiple_files=True, key="s_pdf")

            with col_c:
                st.subheader("📉 Compras y Gastos")
                purch_json = st.file_uploader("Arrastra tus JSON de Compras (Múltiples)", type=["json"], accept_multiple_files=True, key="p_json")
                purch_pdf = st.file_uploader("Arrastra tus PDFs de Compras (Múltiples)", type=["pdf", "zip"], accept_multiple_files=True, key="p_pdf")
                
            st.divider()
            st.subheader("📝 Notas Aclaratorias, Sugerencias y Observaciones por Documento o Mes")
            client_notes = st.text_area(
                "Usa este espacio para detallar aclaraciones sobre documentos específicos:",
                placeholder="Ej. El DTE-03 número... corresponde a una anulación extemporánea...",
                key="notes_input"
            )
                
            submit_files = st.form_submit_button("🚀 Validar y Enviar Documentación con Notas", use_container_width=True)
            
            if submit_files:
                if sales_json or purch_json:
                    json_valido = True
                    archivo_fallido = ""
                    error_detallado = ""
                    
                    for j_file in (sales_json or []) + (purch_json or []):
                        try:
                            j_file.seek(0)
                            content = j_file.read()
                            if isinstance(content, bytes):
                                text_content = content.decode('utf-8-sig', errors='replace')
                            else:
                                text_content = content
                            json.loads(text_content)
                        except Exception as e:
                            json_valido = False
                            archivo_fallido = j_file.name
                            error_detallado = str(e)
                            break
                            
                    if json_valido:
                        s_json_saved = save_files_to_folder(sales_json, st.session_state.username, periodo_str, "sales_json")
                        s_pdf_saved = save_files_to_folder(sales_pdf, st.session_state.username, periodo_str, "sales_pdf")
                        p_json_saved = save_files_to_folder(purch_json, st.session_state.username, periodo_str, "purch_json")
                        p_pdf_saved = save_files_to_folder(purch_pdf, st.session_state.username, periodo_str, "purch_pdf")
                        
                        submission_record = {
                            "user_id": current_user_id,
                            "client": st.session_state.username,
                            "periodo": periodo_str,
                            "sales_json_list": s_json_saved,
                            "sales_pdf_list": s_pdf_saved,
                            "purch_json_list": p_json_saved,
                            "purch_pdf_list": p_pdf_saved,
                            "notes": client_notes,
                            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")
                        }
                        
                        save_submission_to_disk(submission_record)
                        st.success(f"¡Estructura validada! Documentos y notas del periodo {periodo_str} enviados correctamente a RI Consultores.")
                        st.rerun()
                    else:
                        st.error(f"❌ Error en el archivo '{archivo_fallido}': {error_detallado}")
                else:
                    st.warning("Adjunta al menos un archivo JSON principal antes de enviar.")

    with client_tab2:
        st.subheader("💼 MANTENIMIENTO DE PERSONAL Y PLANILLA QUINCENAL")
        
        if current_user_id not in st.session_state.employees_db:
            st.session_state.employees_db[current_user_id] = []
            
        emp_tab_add, emp_tab_manage, emp_tab_calc = st.tabs(["➕ Cargar Empleado", "✏️ Editar / Borrar Empleados", "🧮 Cálculo Masivo de Planilla"])
        
        with emp_tab_add:
            with st.form("form_add_employee"):
                col_e1, col_e2 = st.columns(2)
                with col_e1:
                    new_emp_name = st.text_input("Nombre Completo del Empleado")
                    new_emp_dui = st.text_input("DUI del Empleado (ej. 00000000-0)")
                with col_e2:
                    new_emp_salario = st.number_input("Salario Base Mensual ($)", min_value=0.0, step=10.0, key="add_emp_sal")
                    new_emp_codigo = st.selectbox("Clasificación / Código de Renta", ["Código 01 (Sueldos y Salarios / Tabla)", "Código 60 (Otras Retenciones / Servicios)"], key="add_emp_cod")
                
                submit_add_emp = st.form_submit_button("Cargar Empleado al Sistema", use_container_width=True)
                
                if submit_add_emp:
                    if new_emp_name.strip():
                        if current_user_id not in st.session_state.employees_db:
                            st.session_state.employees_db[current_user_id] = []
                        st.session_state.employees_db[current_user_id].append({
                            "nombre": new_emp_name.strip(),
                            "dui": new_emp_dui.strip(),
                            "codigo": new_emp_codigo,
                            "salario": new_emp_salario
                        })
                        save_json_db(EMPLOYEES_FILE, st.session_state.employees_db)
                        st.success(f"¡Empleado **{new_emp_name.strip()}** cargado al sistema exitosamente con su DUI y código de renta!")
                        st.rerun()
                    else:
                        st.warning("Por favor ingresa el nombre del empleado.")
                        
        with emp_tab_manage:
            st.subheader("Listado de Personal Registrado")
            emps = st.session_state.employees_db.get(current_user_id, [])
            if emps:
                for idx, emp in enumerate(emps):
                    dui_txt = f" — DUI: {emp.get('dui', 'N/A')}" if emp.get('dui') else ""
                    cod_txt = f" | {emp.get('codigo', 'Código 01')}"
                    with st.expander(f"👤 {emp['nombre']}{dui_txt}{cod_txt} — Salario: ${emp['salario']:,.2f}"):
                        with st.form(f"edit_delete_emp_{idx}"):
                            ed_name = st.text_input("Editar Nombre", value=emp['nombre'], key=f"ed_name_{idx}")
                            ed_dui = st.text_input("Editar DUI", value=emp.get('dui', ''), key=f"ed_dui_{idx}")
                            ed_codigo = st.selectbox("Editar Código de Renta", ["Código 01 (Sueldos y Salarios / Tabla)", "Código 60 (Otras Retenciones / Servicios)"], index=0 if "01" in emp.get('codigo', '') else 1, key=f"ed_cod_{idx}")
                            ed_salario = st.number_input("Editar Salario Base Mensual ($)", min_value=0.0, step=10.0, value=float(emp['salario']), key=f"ed_sal_{idx}")
                            
                            col_btn1, col_btn2 = st.columns(2)
                            update_btn = col_btn1.form_submit_button("💾 Guardar Cambios", use_container_width=True)
                            delete_btn = col_btn2.form_submit_button("🗑️ Borrar Empleado", use_container_width=True)
                            
                            if update_btn:
                                st.session_state.employees_db[current_user_id][idx] = {
                                    "nombre": ed_name.strip(),
                                    "dui": ed_dui.strip(),
                                    "codigo": ed_codigo,
                                    "salario": ed_salario
                                }
                                save_json_db(EMPLOYEES_FILE, st.session_state.employees_db)
                                st.success("¡Empleado actualizado y guardado correctamente!")
                                st.rerun()
                                
                            if delete_btn:
                                st.session_state.employees_db[current_user_id].pop(idx)
                                save_json_db(EMPLOYEES_FILE, st.session_state.employees_db)
                                st.success("¡Empleado eliminado del sistema permanentemente!")
                                st.rerun()
            else:
                st.info("No hay empleados cargados en el sistema todavía.")
                
        with emp_tab_calc:
            st.markdown(f"##### 🧮 Cálculo Masivo de Planilla Quincenal — **{periodo_str}** ({quincena_op})")
            st.info("ℹ️ Aquí puedes ingresar las bonificaciones u otros ingresos extraordinarios para cada empleado. El sistema calculará automáticamente el sueldo quincenal, ISSS, AFP y Renta para todo el personal de forma simultánea.")
            
            emps_calc = st.session_state.employees_db.get(current_user_id, [])
            
            if emps_calc:
                with st.form("form_calc_all_employees_quincenal"):
                    updated_emp_inputs = []
                    for idx, emp in enumerate(emps_calc):
                        st.markdown(f"**👤 {emp['nombre']}** (DUI: {emp.get('dui', 'N/A')} | {emp.get('codigo', 'Código 01')})")
                        col_q1, col_q2 = st.columns(2)
                        with col_q1:
                            salario_base_mensual = emp['salario']
                            st.text(f"Salario Base Mensual: ${salario_base_mensual:,.2f}")
                        with col_q2:
                            bono_val = st.number_input(
                                f"Bonificaciones / Otros Ingresos ($)", 
                                min_value=0.0, 
                                value=0.0, 
                                step=5.0, 
                                key=f"q_bono_{idx}"
                            )
                        st.divider()
                        updated_emp_inputs.append({
                            "nombre": emp['nombre'],
                            "dui": emp.get('dui', 'N/A'),
                            "codigo": emp.get('codigo', 'Código 01'),
                            "salario_mensual": salario_base_mensual,
                            "bonificacion": bono_val
                        })
                    
                    btn_calc_all_emps = st.form_submit_button("🚀 Calcular Planilla Preliminar Completa", use_container_width=True)
                    
                if btn_calc_all_emps:
                    planilla_q_data = []
                    tot_devengado = 0.0
                    tot_isss = 0.0
                    tot_afp = 0.0
                    tot_renta = 0.0
                    tot_liquido = 0.0
                    
                    for item in updated_emp_inputs:
                        sueldo_base_q = item['salario_mensual'] / 2.0
                        devengado_q = sueldo_base_q + item['bonificacion']
                        
                        isss_q = 0.0
                        afp_q = 0.0
                        renta_q = 0.0
                        
                        if "01" in item['codigo']:
                            isss_q = calcular_isss_quincenal(devengado_q)
                            afp_q = calcular_afp_quincenal(devengado_q)
                            neto_gravable = devengado_q - isss_q - afp_q
                            renta_q = calcular_renta_quincenal(neto_gravable)
                        else: # Código 60
                            renta_q = round(devengado_q * 0.10, 2)
                            
                        liquido_q = round(devengado_q - isss_q - afp_q - renta_q, 2)
                        
                        tot_devengado += devengado_q
                        tot_isss += isss_q
                        tot_afp += afp_q
                        tot_renta += renta_q
                        tot_liquido += liquido_q
                        
                        planilla_q_data.append({
                            "Empleado": item['nombre'],
                            "DUI": item['dui'],
                            "Código Renta": item['codigo'],
                            "Salario Base M.": item['salario_mensual'],
                            "Bonif. / Ingresos": item['bonificacion'],
                            "Total Devengado Q.": devengado_q,
                            "ISSS": isss_q,
                            "AFP": afp_q,
                            "RENTA": renta_q,
                            "Líquido a Pagar": liquido_q
                        })
                    
                    st.session_state["temp_planilla_quincenal"] = {
                        "periodo": periodo_str,
                        "quincena": quincena_op,
                        "data": planilla_q_data,
                        "total_devengado": tot_devengado,
                        "total_isss": tot_isss,
                        "total_afp": tot_afp,
                        "total_renta": tot_renta,
                        "total_liquido": tot_liquido
                    }
                    st.success("¡Planilla quincenal preliminar calculada con éxito para todo el personal!")
                    st.rerun()
                
                if "temp_planilla_quincenal" in st.session_state:
                    pq_info = st.session_state["temp_planilla_quincenal"]
                    if pq_info["periodo"] == periodo_str:
                        st.markdown("---")
                        st.markdown("### 📋 Vista Preliminar de Planilla Quincenal")
                        df_pq = pd.DataFrame(pq_info["data"])
                        
                        col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
                        col_m1.metric("Total Devengado", f"${pq_info['total_devengado']:,.2f}")
                        col_m2.metric("Total ISSS", f"${pq_info['total_isss']:,.2f}")
                        col_m3.metric("Total AFP", f"${pq_info['total_afp']:,.2f}")
                        col_m4.metric("Total RENTA", f"${pq_info['total_renta']:,.2f}")
                        col_m5.metric("Líquido a Pagar", f"${pq_info['total_liquido']:,.2f}")
                        
                        st.dataframe(
                            df_pq.style.format({
                                "Salario Base M.": "${:,.2f}",
                                "Bonif. / Ingresos": "${:,.2f}",
                                "Total Devengado Q.": "${:,.2f}",
                                "ISSS": "${:,.2f}",
                                "AFP": "${:,.2f}",
                                "RENTA": "${:,.2f}",
                                "Líquido a Pagar": "${:,.2f}"
                            }),
                            use_container_width=True
                        )
                        
                        # Generador de Excel Profesional para descarga
                        output_pq = io.BytesIO()
                        total_row_pq = pd.DataFrame([{
                            "Empleado": "TOTALES GENERALES",
                            "DUI": "",
                            "Código Renta": "",
                            "Salario Base M.": 0.0,
                            "Bonif. / Ingresos": 0.0,
                            "Total Devengado Q.": pq_info['total_devengado'],
                            "ISSS": pq_info['total_isss'],
                            "AFP": pq_info['total_afp'],
                            "RENTA": pq_info['total_renta'],
                            "Líquido a Pagar": pq_info['total_liquido']
                        }])
                        df_exp_pq = pd.concat([df_pq, total_row_pq], ignore_index=True)
                        
                        with pd.ExcelWriter(output_pq, engine='openpyxl') as writer:
                            df_exp_pq.to_excel(writer, index=False, sheet_name='Planilla Quincenal', startrow=4)
                            workbook = writer.book
                            worksheet = writer.sheets['Planilla Quincenal']
                            
                            worksheet['A1'] = f"RI CONSULTORES — PLANILLA DE SUELDOS QUINCENAL ({quincena_op.upper()})"
                            worksheet['A2'] = f"Periodo Fiscal: {periodo_str}"
                            worksheet['A3'] = f"Cliente / Contribuyente: {st.session_state.username}"
                            
                            title_font = Font(name='Calibri', size=12, bold=True, color='1F4E78')
                            subtitle_font = Font(name='Calibri', size=10, bold=True, color='595959')
                            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                            total_font = Font(name='Calibri', size=11, bold=True, color='000000')
                            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                            
                            align_center = Alignment(horizontal='center', vertical='center')
                            align_right = Alignment(horizontal='right', vertical='center')
                            align_left = Alignment(horizontal='left', vertical='center')
                            
                            thin_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
                            )
                            thick_bottom_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='double', color='000000')
                            )
                            
                            worksheet['A1'].font = title_font
                            worksheet['A2'].font = subtitle_font
                            worksheet['A3'].font = subtitle_font
                            
                            header_row_idx = 5
                            for col_num in range(1, len(df_exp_pq.columns) + 1):
                                cell = worksheet.cell(row=header_row_idx, column=col_num)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = align_center
                                cell.border = thin_border
                                
                            total_row_idx = header_row_idx + len(df_exp_pq)
                            for row_idx in range(header_row_idx + 1, total_row_idx + 1):
                                is_total_row = (row_idx == total_row_idx)
                                for col_idx in range(1, len(df_exp_pq.columns) + 1):
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    if is_total_row:
                                        cell.font = total_font
                                        cell.fill = total_fill
                                        cell.border = thick_bottom_border
                                    else:
                                        cell.border = thin_border
                                        
                                    if col_idx >= 4:
                                        cell.number_format = '$#,##0.00'
                                        cell.alignment = align_right
                                    elif col_idx in [2, 3]:
                                        cell.alignment = align_center
                                    else:
                                        cell.alignment = align_left
                                        
                            for col in worksheet.columns:
                                max_len = 0
                                for cell in col:
                                    if cell.row >= 5:
                                        val_str = str(cell.value or '')
                                        if len(val_str) > max_len:
                                            max_len = len(val_str)
                                col_letter = get_column_letter(col[0].column)
                                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 18)
                                
                        excel_pq_bytes = output_pq.getvalue()
                        safe_c_name = st.session_state.username.replace(" ", "_").replace(".", "")
                        
                        st.download_button(
                            label="📥 Descargar Planilla Quincenal en Excel (Formato Profesional)",
                            data=excel_pq_bytes,
                            file_name=f"Planilla_Quincenal_{safe_c_name}_{periodo_str.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        col_act1, col_act2 = st.columns(2)
                        with col_act1:
                            if st.button("📤 Enviar Planilla Oficial a RI Consultores", use_container_width=True, type="primary"):
                                submission_pq_record = {
                                    "user_id": current_user_id,
                                    "client": st.session_state.username,
                                    "periodo": periodo_str,
                                    "quincena": quincena_op,
                                    "type": "planilla_quincenal_empleados",
                                    "data": pq_info["data"],
                                    "total_devengado": pq_info["total_devengado"],
                                    "total_isss": pq_info["total_isss"],
                                    "total_afp": pq_info["total_afp"],
                                    "total_renta": pq_info["total_renta"],
                                    "total_liquido": pq_info["total_liquido"],
                                    "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")
                                }
                                save_submission_to_disk(submission_pq_record)
                                st.success(f"¡Planilla quincenal del periodo {periodo_str} enviada exitosamente al administrador!")
                        with col_act2:
                            if st.button("🗑️ Borrar / Descartar Cálculo Preliminar", use_container_width=True):
                                del st.session_state["temp_planilla_quincenal"]
                                st.rerun()
            else:
                st.warning("⚠️ No hay empleados registrados en el sistema. Ve a la pestaña **'➕ Cargar Empleado'** para agregar personal antes de calcular la planilla.")

    with client_tab3:
        st.subheader("📊 Historial de Declaraciones, Resumen Ejecutivo de IVA y Trazabilidad DTE")
        
        if mis_envios:
            st.info("Visualiza el detalle de tus declaraciones, las notas enviadas, el resumen de IVA y los códigos de generación correspondientes.")
            for envio in mis_envios:
                if envio.get("type") in ["planilla_eventuales", "planilla_quincenal_empleados"]:
                    continue
                with st.expander(f"📅 Periodo: {envio['periodo']} — Entregado el {envio['fecha']}"):
                    if envio.get('notes'):
                        st.markdown("##### 📝 Tus Notas / Aclaraciones Enviadas")
                        st.info(envio['notes'])
                        st.markdown("---")
                    
                    df_sales = extract_invoice_summary(envio.get('sales_json_list'))
                    df_purch = extract_invoice_summary(envio.get('purch_json_list'))
                    
                    v_val = df_sales['Valor'].sum() if not df_sales.empty else 0.0
                    v_iva = df_sales['IVA'].sum() if not df_sales.empty else 0.0
                    v_tot = df_sales['Total'].sum() if not df_sales.empty else 0.0
                    
                    p_val = df_purch['Valor'].sum() if not df_purch.empty else 0.0
                    p_iva = df_purch['IVA'].sum() if not df_purch.empty else 0.0
                    p_tot = df_purch['Total'].sum() if not df_purch.empty else 0.0
                    
                    st.markdown("##### 💼 Resumen Ejecutivo de IVA")
                    col_re1, col_re2, col_re3 = st.columns(3)
                    col_re1.metric("Débito Fiscal (IVA Ventas)", f"${v_iva:,.2f}")
                    col_re2.metric("Crédito Fiscal (IVA Compras)", f"${p_iva:,.2f}")
                    
                    iva_neto = v_iva - p_iva
                    if iva_neto >= 0:
                        col_re3.metric("IVA a Pagar Estimado", f"${iva_neto:,.2f}", delta_color="inverse")
                    else:
                        col_re3.metric("Remanente de IVA a Favor", f"${abs(iva_neto):,.2f}", delta_color="normal")
                    
                    st.markdown("---")
                    st.markdown("##### 📈 Detalle de Ventas (Trazabilidad DTE)")
                    if not df_sales.empty:
                        col_m1, col_m2, col_m3 = st.columns(3)
                        col_m1.metric("Subtotal Ventas", f"${v_val:,.2f}")
                        col_m2.metric("IVA Ventas", f"${v_iva:,.2f}")
                        col_m3.metric("Total Ventas", f"${v_tot:,.2f}")
                        
                        st.dataframe(
                            df_sales.style.format({
                                "Valor": "${:,.2f}",
                                "IVA": "${:,.2f}",
                                "Total": "${:,.2f}"
                            }),
                            use_container_width=True
                        )
                    else:
                        st.text("Sin registros de ventas detallados para este periodo.")
                        
                    st.markdown("---")
                    st.markdown("##### 📉 Detalle de Compras y Gastos (Trazabilidad DTE)")
                    if not df_purch.empty:
                        col_pm1, col_pm2, col_pm3 = st.columns(3)
                        col_pm1.metric("Subtotal Compras", f"${p_val:,.2f}")
                        col_pm2.metric("IVA Compras", f"${p_iva:,.2f}")
                        col_pm3.metric("Total Compras", f"${p_tot:,.2f}")
                        
                        st.dataframe(
                            df_purch.style.format({
                                "Valor": "${:,.2f}",
                                "IVA": "${:,.2f}",
                                "Total": "${:,.2f}"
                            }),
                            use_container_width=True
                        )
                    else:
                        st.text("Sin registros de compras detallados para este periodo.")
        else:
            st.warning("⚠️ Aún no has registrado envíos de documentos en el portal.")

    with client_tab4:
        st.subheader("🧾 BASE MAESTRO Y PLANILLA DE RETENCIÓN DE RENTA (10% EVENTUALES)")
        st.markdown("Administra aquí tu **directorio permanente** de prestadores eventuales. Los datos que guardes aquí se mantendrán fijos y se precargarán automáticamente cada mes al generar tus planillas.")
        
        if current_user_id not in st.session_state.eventuales_db:
            st.session_state.eventuales_db[current_user_id] = []
            
        ev_tab_master, ev_tab_calc = st.tabs(["📂 Directorio Maestro (Base Fija)", "🧮 Generar Planilla del Periodo Actual"])
        
        with ev_tab_master:
            st.markdown("##### ➕ Registrar Nuevo Prestador en la Base Maestro")
            with st.form("form_add_eventual_master"):
                col_m1, col_m2, col_m3 = st.columns(3)
                with col_m1:
                    new_ev_name = st.text_input("Nombre Completo")
                with col_m2:
                    new_ev_dui = st.text_input("DUI (ej. 00000000-0)")
                with col_m3:
                    new_ev_monto = st.number_input("Monto Bruto Habitual ($)", min_value=0.0, step=10.0, value=0.0)
                    
                submit_add_ev = st.form_submit_button("💾 Guardar en la Base Maestro", use_container_width=True)
                
                if submit_add_ev:
                    if new_ev_name.strip():
                        if current_user_id not in st.session_state.eventuales_db:
                            st.session_state.eventuales_db[current_user_id] = []
                        st.session_state.eventuales_db[current_user_id].append({
                            "nombre": new_ev_name.strip(),
                            "dui": new_ev_dui.strip(),
                            "monto": new_ev_monto
                        })
                        save_json_db(EVENTUALES_FILE, st.session_state.eventuales_db)
                        st.success(f"¡Prestador **{new_ev_name.strip()}** guardado permanentemente en la base maestro!")
                        st.rerun()
                    else:
                        st.warning("Por favor ingresa al menos el nombre del prestador.")
                        
            st.divider()
            st.markdown("##### ✏️ Listado Fijo Actual (Editar o Eliminar de la Base)")
            evs_master = st.session_state.eventuales_db.get(current_user_id, [])
            if evs_master:
                for idx, ev in enumerate(evs_master):
                    dui_str = f" — DUI: {ev['dui']}" if ev['dui'] else " — Sin DUI"
                    with st.expander(f"👤 {ev['nombre']}{dui_str} | Base: ${ev['monto']:,.2f}"):
                        with st.form(f"edit_delete_ev_master_{idx}"):
                            ed_ev_name = st.text_input("Editar Nombre", value=ev['nombre'], key=f"ed_ev_name_{idx}")
                            ed_ev_dui = st.text_input("Editar DUI", value=ev['dui'], key=f"ed_ev_dui_{idx}")
                            ed_ev_monto = st.number_input("Editar Monto Bruto Base ($)", min_value=0.0, step=10.0, value=float(ev['monto']), key=f"ed_ev_mon_{idx}")
                            
                            col_btn1, col_btn2 = st.columns(2)
                            update_ev_btn = col_btn1.form_submit_button("💾 Actualizar Maestro", use_container_width=True)
                            delete_ev_btn = col_btn2.form_submit_button("🗑️ Eliminar de la Base", use_container_width=True)
                            
                            if update_ev_btn:
                                st.session_state.eventuales_db[current_user_id][idx] = {
                                    "nombre": ed_ev_name.strip(),
                                    "dui": ed_ev_dui.strip(),
                                    "monto": ed_ev_monto
                                }
                                save_json_db(EVENTUALES_FILE, st.session_state.eventuales_db)
                                st.success("¡Base maestro actualizada correctamente!")
                                st.rerun()
                                
                            if delete_ev_btn:
                                st.session_state.eventuales_db[current_user_id].pop(idx)
                                save_json_db(EVENTUALES_FILE, st.session_state.eventuales_db)
                                st.success("¡Prestador eliminado permanentemente de la base!")
                                st.rerun()
            else:
                st.info("Tu base maestro de eventuales está vacía. Agrega prestadores arriba para que aparezcan fijos cada mes.")
                
        with ev_tab_calc:
            st.markdown(f"##### 🧮 Planilla de Retención para el Periodo Seleccionado: **{periodo_str}**")
            st.info("ℹ️ Los prestadores registrados en tu **Base Maestro** se cargan automáticamente abajo. Puedes confirmar o ajustar el monto bruto devengado en este periodo específico antes de exportar.")
            
            evs_calc = st.session_state.eventuales_db.get(current_user_id, [])
            
            if evs_calc:
                with st.form("form_calc_all_eventuales"):
                    updated_evs_data = []
                    for idx, ev in enumerate(evs_calc):
                        col_n, col_d, col_m = st.columns([2, 1, 1])
                        with col_n:
                            st.text(ev['nombre'])
                        with col_d:
                            st.text(ev['dui'] or 'Sin DUI')
                        with col_m:
                            monto_val = st.number_input(
                                f"Monto ($) - {ev['nombre']}", 
                                min_value=0.0, 
                                value=float(ev['monto']), 
                                step=10.0, 
                                key=f"calc_all_monto_{idx}",
                                label_visibility="collapsed"
                            )
                        updated_evs_data.append({
                            "nombre": ev['nombre'],
                            "dui": ev['dui'],
                            "monto": monto_val
                        })
                    
                    btn_calc_all = st.form_submit_button("🚀 Calcular Planilla de Este Periodo", use_container_width=True)
                    
                if btn_calc_all:
                    planilla_data = []
                    total_bruto = 0.0
                    total_retencion = 0.0
                    total_liquido = 0.0
                    
                    for item in updated_evs_data:
                        bruto = item['monto']
                        retencion = bruto * 0.10
                        liquido = bruto - retencion
                        
                        total_bruto += bruto
                        total_retencion += retencion
                        total_liquido += liquido
                        
                        planilla_data.append({
                            "Prestador Eventual": item['nombre'],
                            "DUI": item['dui'] or "N/A",
                            "Monto Bruto": bruto,
                            "Retención 10%": retencion,
                            "Líquido a Pagar": liquido
                        })
                    
                    st.session_state["temp_planilla_eventuales"] = {
                        "periodo": periodo_str,
                        "data": planilla_data,
                        "total_bruto": total_bruto,
                        "total_retencion": total_retencion,
                        "total_liquido": total_liquido
                    }
                    st.success(f"¡Planilla de Eventuales para el periodo **{periodo_str}** calculada con éxito!")
                
                if "temp_planilla_eventuales" in st.session_state:
                    p_info = st.session_state["temp_planilla_eventuales"]
                    if p_info["periodo"] == periodo_str:
                        df_planilla = pd.DataFrame(p_info["data"])
                        total_bruto = p_info["total_bruto"]
                        total_retencion = p_info["total_retencion"]
                        total_liquido = p_info["total_liquido"]
                        
                        st.markdown("##### 📊 Resumen Consolidado")
                        col_m1, col_m2, col_m3 = st.columns(3)
                        col_m1.metric("Total Bruto", f"${total_bruto:,.2f}")
                        col_m2.metric("Total Retención 10%", f"${total_retencion:,.2f}")
                        col_m3.metric("Total Líquido a Pagar", f"${total_liquido:,.2f}")
                        
                        st.markdown("##### 📋 Detalle de la Planilla")
                        st.dataframe(
                            df_planilla.style.format({
                                "Monto Bruto": "${:,.2f}",
                                "Retención 10%": "${:,.2f}",
                                "Líquido a Pagar": "${:,.2f}"
                            }),
                            use_container_width=True
                        )
                        
                        output = io.BytesIO()
                        total_row_df = pd.DataFrame([{
                            "Prestador Eventual": "TOTALES GENERALES",
                            "DUI": "",
                            "Monto Bruto": total_bruto,
                            "Retención 10%": total_retencion,
                            "Líquido a Pagar": total_liquido
                        }])
                        df_export = pd.concat([df_planilla, total_row_df], ignore_index=True)
                        
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_export.to_excel(writer, index=False, sheet_name='Planilla Eventuales', startrow=4)
                            workbook = writer.book
                            worksheet = writer.sheets['Planilla Eventuales']
                            
                            worksheet['A1'] = "RI CONSULTORES — PLANILLA DE RETENCIÓN DE RENTA (10% EVENTUALES)"
                            worksheet['A2'] = f"Periodo Fiscal / Mes de Referencia: {periodo_str}"
                            worksheet['A3'] = f"Cliente / Contribuyente: {st.session_state.username}"
                            
                            title_font = Font(name='Calibri', size=12, bold=True, color='1F4E78')
                            subtitle_font = Font(name='Calibri', size=10, bold=True, color='595959')
                            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                            total_font = Font(name='Calibri', size=11, bold=True, color='000000')
                            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                            
                            align_center = Alignment(horizontal='center', vertical='center')
                            align_right = Alignment(horizontal='right', vertical='center')
                            align_left = Alignment(horizontal='left', vertical='center')
                            
                            thin_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
                            )
                            thick_bottom_border = Border(
                                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='double', color='000000')
                            )
                            
                            worksheet['A1'].font = title_font
                            worksheet['A2'].font = subtitle_font
                            worksheet['A3'].font = subtitle_font
                            
                            header_row_idx = 5
                            for col_num in range(1, len(df_export.columns) + 1):
                                cell = worksheet.cell(row=header_row_idx, column=col_num)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = align_center
                                cell.border = thin_border
                                
                            total_row_idx = header_row_idx + len(df_export)
                            for row_idx in range(header_row_idx + 1, total_row_idx + 1):
                                is_total_row = (row_idx == total_row_idx)
                                for col_idx in range(1, len(df_export.columns) + 1):
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    if is_total_row:
                                        cell.font = total_font
                                        cell.fill = total_fill
                                        cell.border = thick_bottom_border
                                    else:
                                        cell.border = thin_border
                                        
                                    if col_idx in [3, 4, 5]:
                                        cell.number_format = '$#,##0.00'
                                        cell.alignment = align_right
                                    elif col_idx == 2:
                                        cell.alignment = align_center
                                    else:
                                        cell.alignment = align_left
                                        
                            for col in worksheet.columns:
                                max_len = 0
                                for cell in col:
                                    if cell.row >= 5:
                                        val_str = str(cell.value or '')
                                        if len(val_str) > max_len:
                                            max_len = len(val_str)
                                col_letter = get_column_letter(col[0].column)
                                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 18)
                                
                        excel_data = output.getvalue()
                        safe_client_name = st.session_state.username.replace(" ", "_").replace(".", "")
                        file_name_download = f"Planilla_Eventuales_10_{safe_client_name}_{periodo_str.replace(' ', '_')}.xlsx"
                        
                        st.download_button(
                            label="📥 Descargar Planilla de Eventuales en Excel (Formato Profesional)",
                            data=excel_data,
                            file_name=file_name_download,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        st.divider()
                        if st.button("📤 Enviar Planilla Oficial de Eventuales a RI Consultores", use_container_width=True, type="primary"):
                            planilla_submission_record = {
                                "user_id": current_user_id,
                                "client": st.session_state.username,
                                "periodo": periodo_str,
                                "type": "planilla_eventuales",
                                "data": p_info["data"],
                                "total_bruto": total_bruto,
                                "total_retencion": total_retencion,
                                "total_liquido": total_liquido,
                                "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")
                            }
                            save_submission_to_disk(planilla_submission_record)
                            st.success(f"¡Planilla de eventuales del periodo {periodo_str} enviada exitosamente al administrador!")
            else:
                st.warning("⚠️ Tu Base Maestro está vacía. Ve a la pestaña **'Directorio Maestro (Base Fija)'** para registrar a tus prestadores eventuales por primera vez.")

    with client_tab5:
        auditor_deliverables_portal(st.session_state.username)

# --- Control de Sesión y Redirección Principal ---
if not st.session_state.logged_in:
    login_screen()
else:
    with st.sidebar:
        st.write(f"Conectado como:\n**{st.session_state.username}**")
        st.divider()
        if st.button("Cerrar Sesión", type="primary"):
            st.session_state.logged_in = False
            st.session_state.user_role = None
            st.session_state.username = ""
            st.session_state.user_id = ""
            st.rerun()
            
    if st.session_state.user_role == "admin":
        admin_dashboard()
    elif st.session_state.user_role == "client":
        client_dashboard()
